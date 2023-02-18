import selenium.common.exceptions
import time
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
# from selenium.common.exceptions import NoSuchElementException
# from selenium.common.exceptions import StaleElementReferenceException
# from selenium.webdriver.support import expected_conditions
# from selenium.webdriver.support.ui import WebDriverWait
# from selenium.webdriver.support import expected_conditions as EC

from getpass import getpass
import openpyxl as xl


user_id = input("Enter User ID: ")
user_password = getpass()

PATH = "C:\Program Files (x86)\chromedriver.exe"
chrome_options = Options() 
chrome_options.add_experimental_option("detach", True)

driver = webdriver.Chrome(options=chrome_options)

wb = xl.load_workbook("List of Files for Scaning.xlsx", data_only=True)
sheet = wb.worksheets[0]

driver.get("http://172.16.0.6:81/ords/f?p=700:LOGIN:5055859720039:::::")

user_name = driver.find_element(by=By.ID, value="P101_USERNAME")
user_name.send_keys(user_id)

password = driver.find_element(by=By.ID, value="P101_PASSWORD")
password.send_keys(user_password)

general_key = driver.find_element(by=By.ID, value="P101_LOGIN_CODE")
general_key.send_keys("123456")

sign_in = driver.find_element(by=By.ID, value="P101_LOGIN")
sign_in.click()

exam = driver.find_element(by=By.CLASS_NAME, value ="t-Tabs-label")
exam.click()


for row in range(1, sheet.max_row + 1):
    cell = sheet.cell(row, 2)
    if cell.value is not None:        
        student = cell.value
        check_id = student
        stu_id = driver.find_element(by=By.ID, value="P2790_V_DIRECT_STUDENT_ID")
        count = 0
        # wait for element "P2790_V_DIRECT_STUDENT_ID" to be loaded on page
        while count < 3:
            count += 1
            try:                
                stu_id.clear()
                stu_id.send_keys(check_id)
                stu_id.send_keys(Keys.ENTER)
                break
            except Exception as e:
                time.sleep(1)
        try:
            # if "t-Report-report" is not on page means there is no degree data. Degree is not made.
            report = driver.find_element(by=By.CLASS_NAME, value ="t-Report-report")
            working = report.text.split("\n")[1].split("\t")[0]
            raw = working.split(" ")
            date, degree_no = raw[0], raw[1]
            status = working.split("   ")[1]
            excel_date, excel_degree_no, excel_status = sheet.cell(row, 9), sheet.cell(row, 10), sheet.cell(row, 11)
            excel_date.value = date
            excel_degree_no.value = degree_no
            excel_status.value = status
            wb.save("List of Files for Scaning.xlsx")
            print(check_id.strip(), date, degree_no, status)            
        except selenium.common.exceptions.NoSuchElementException:
            print(check_id.strip() + " - Degree No Issued")
driver.close()