from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys


import selenium.common.exceptions
from selenium.webdriver.support.ui import Select
from selenium.webdriver.support.wait import WebDriverWait
from selenium.common.exceptions import NoSuchElementException
from selenium.common.exceptions import StaleElementReferenceException
from selenium.webdriver.support import expected_conditions
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

import time
import openpyxl as xl


# Execute your JavaScript code on the page
js_code = """
var get_ready = false
var opening_balance = 0
var debit = 0
var credit = 0
var data = [];
var table = document.getElementById("report_table_R312545805814404689");
for (var i = 1, row; row = table.rows[i]; i++) {    
    for (var j = 0, col; col = row.cells[j]; j++) {
        if (!get_ready) {
            if (row.cells[2].innerHTML == 'Fall-2023') {
                 var previousRow = table.rows[i - 1];
                 removing_comas = previousRow.cells[7].innerHTML.replace(",", "");
                 opening_balance = parseInt(removing_comas);             
                 get_ready = true
            }
        }
        if (get_ready) {
            if (row.cells[5].innerHTML != "" && !row.cells[5].innerHTML.includes("strong")) {
                removing_comas = row.cells[5].innerHTML.replace(",", "");
                debit += parseInt(removing_comas) / 8;             
            }
            if (row.cells[6].innerHTML != "" && !row.cells[6].innerHTML.includes("strong")) {
                removing_comas = row.cells[6].innerHTML.replace(",", "");
                credit += parseInt(removing_comas) / 8;  
            }
        }
    }
}

data.push(opening_balance)
data.push(debit)
data.push(credit)
data.push(opening_balance + debit - credit)

return data;
"""



# Don't forget to close the WebDriver when done




user_id = input("Enter User: ")
user_password = input("Password: ")


chrome_options = Options()
chrome_options.add_argument("--disable-popup-blocking")
chrome_options.add_experimental_option("detach", True)
driver = webdriver.Chrome(options=chrome_options)

wb = xl.load_workbook("Students_List.xlsx", data_only=True)
sheet = wb.worksheets[0]

URL = "http://faculty.induscms.com:81/ords/r/erasoft/a500550500/login_desktop"

driver.get(URL)

user_name = driver.find_element(by=By.ID, value="P101_USERNAME")
user_name.send_keys(user_id)

password = driver.find_element(by=By.ID, value="P101_PASSWORD")
password.send_keys(user_password)

general_key = driver.find_element(by=By.ID, value="P101_LOGIN_CODE")
general_key.send_keys("123456")

sign_in = driver.find_element(by=By.ID, value="LOGIN")
sign_in.click()

click_menu = driver.find_element(by=By.ID, value="t_Button_navControl")
count = 0
# wait for element "P2790_V_DIRECT_STUDENT_ID" to be loaded on page

ready_to_loop = False

while count < 3:
    count += 1
    try:
        click_menu.click()
        dropdown_element = driver.find_element(by=By.XPATH, value="//span[@class='t-MegaMenu-label' and text()='Issuing Individual Student Fee Voucher']")
        dropdown_element.click()
        count = 0
        ready_to_loop = True
        break
    except Exception as e:
        time.sleep(1)


if ready_to_loop:
    for row in range(2, sheet.max_row + 1):
        student_id = sheet.cell(row, 1)
        stu_id_element = driver.find_element(by=By.ID, value="P0_V_DIRECT_STUDENT_ID")
        while count < 3:
            count += 1
            try:
                stu_id_element.clear()
                stu_id_element.send_keys(student_id.value)
                stu_id_element.send_keys(Keys.ENTER)
                try:
                    result = driver.execute_script(js_code)
                    print(student_id.value + ": " + str(result))
                    opening_balance_value = result[0]
                    debit_value = result[1]
                    credit_value = result[2]
                    balance = result[3]
                    opening_balance_location, debit_location, credit_location, balance_location = sheet.cell(row, 10), sheet.cell(row, 11), sheet.cell(row, 12), sheet.cell(row, 13)
                    opening_balance_location.value = opening_balance_value
                    debit_location.value = debit_value
                    credit_location.value = credit_value
                    balance_location.value = balance
                    wb.save("Students_List.xlsx")
                except Exception as e:
                    continue
                count = 0
                break
            except Exception as e:
                time.sleep(1)
